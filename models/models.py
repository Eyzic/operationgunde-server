from openpyxl import load_workbook
from datetime import datetime

#{'user_id', 'date', 'hrv', 'sleeping_hours', 'stress_level', 'muscle_ache', 'mood_level', 'injury_level', 'energy_level'}
#Tidstämpel	Namn:	HRV	Hur många timmar sov du i natt? 	Stressnivå 	Muskelvärk & muskeltrötthet	Humör 	Har du några skador? Ange graden på skadorna	Hur är din energinivå just nu?	Togs mätningen direkt efter du vaknade upp?  	RMSSD	Ln(RMSSD)	SDNN

def add_to_hrv_xlsx(stats):

    filename = "models/training.xlsx"

    date = datetime.fromisoformat(stats['date'])
    date = date.strftime("%d/%m/%Y")

    if stats['user_id'] == 2:
        user = 'Filip Helmroth'
    
    new_row = [date, user, stats['hrv'], stats['sleeping_hours'], stats['stress_level'], stats['muscle_ache'], stats['mood_level'], stats['injury_level'], stats['energy_level'], "Ja"]

    print(new_row)
    # Confirm file exists. 
    # If not, create it, add headers, then append new data
    
    wb = load_workbook(filename)
    ws = wb.worksheets[0]  

    ws.append(new_row)
    wb.save(filename)

#{'user_id', 'date', 'training_intensity', 'training_type', 'elapsed_time', 'energy_level'}
#Tidstämpel	Namn:	Träningsdagen	Hur intensiv kände du att träningen var?	Hur länge varade träningspasset (i min):	Typ av träning:	Hur är din energinivå just nu?	Träningsbelastning

def add_to_training_xlsx(stats):
    
    filename = "models/training.xlsx"

    curr_date = datetime.fromisoformat(stats['date'])
    date = curr_date.strftime("%d/%m/%Y")
    day = curr_date.strftime("%A")

    if day == "Monday":
        day = "Måndag"
    elif day == "Tuesday":
        day = "Tisdag"
    elif day == "Wednesday":
        day = "Onsdag"
    elif day == "Thursday":
        day = "Torsdag"
    elif day == "Friday":
        day = "Fredag"
    elif day == "Saturday":
        day = "Lördag"
    elif day == "Sunday":
        day = "Söndag"

    if stats['user_id'] == 2:
        user = 'Filip Helmroth'
    
    new_row = [date, user, day, stats['training_intensity'], stats['elapsed_time'], stats['training_type'], stats['energy_level'], ""]

    print(new_row)
    # Confirm file exists. 
    # If not, create it, add headers, then append new data
    
    wb = load_workbook(filename)
    ws = wb.worksheets[0]  

    ws.append(new_row)
    wb.save(filename)